import os
import datetime
import time
from datetime import timedelta
import random
from transfer.models import (PCTransfer, PalletCellTransfer, 
                            TransferTransferAttribute)
from batch.models import PalletPalletAttribute
from stock.models import Cell, CellLog
from product.models import Product, ProductLog
from django.db.models import Count
from django.db.models import Max

from GlobalVariables import *
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import (Font, Alignment, DEFAULT_FONT)
from openpyxl.worksheet import page
from openpyxl.styles.borders import Border, Side
from files.models import ProductFiles
import pprint

def has_permission(user, permitted_name):
    groups = user.groups.values_list('name',flat=True)
    groups_as_list = list(groups)
    if permitted_name in groups_as_list:
        return True
    else:
        return False

def validate_date(date):
    try:
        datetime.date.fromisoformat(date)
        return True
    except:
        return False


def create_name():
    chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789123"
    char_length = len(chars)
    file_name = ""
    for i in range(12):
        file_name = file_name + chars[random.randrange(char_length)]
    
    return file_name

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def transfer_excel(transfer_id):
    pass

MONTHS = {
    1: "Ýanwar",
    2: "Fewral",
    3: "Mart",
    4: "Aprel",
    5: "Maý",
    6: "Iýun",
    7: "Iýul",
    8: "Awgust",
    9: "Sentýabr",
    10: "Oktýabr",
    11: "Noýabr",
    12: "Dekabr",
}

def talapnama_create(pk):
    try:
        starting = time.perf_counter()
        template_file = 'xls_templates/talapnama.xlsx'
        wb = load_workbook(template_file)
        ws = wb.active
        transfer = PCTransfer.objects.get(pk=pk)
        transfer_created = transfer.created_at
        transfer_date = transfer_created.strftime("%d-%m-%Y")
        current_year = transfer_created.year
        month = transfer_created.month
        current_month = MONTHS[month]
        current_day = transfer_created.day

        transfer_pallets = PalletCellTransfer.objects.filter(
            transfer_id=transfer
        )
        transfer_attributes = TransferTransferAttribute.objects.filter(
                pctransfer_id = transfer
        )
        account_number = ""
        from_whom = ""
        to_whom = ""
        from_whom_dep = ""
        to_whom_dep = ""
        car_model = ""
        car_number = ""
        car_driver = ""
        for attr in transfer_attributes:
            if attr.pctransfer_attribute_id.title == HASAP_NO:
                account_number = attr.value
            if attr.pctransfer_attribute_id.title == TO_WHOM:
                to_whom = attr.value
            if attr.pctransfer_attribute_id.title == FROM_WHOM:
                from_whom = attr.value
            if attr.pctransfer_attribute_id.title == FROM_WHOM_DEP:
                from_whom_dep = attr.value
            if attr.pctransfer_attribute_id.title == TO_WHOM_DEP:
                to_whom_dep = attr.value
            if attr.pctransfer_attribute_id.title == CAR_MODEL:
                car_model = attr.value
            if attr.pctransfer_attribute_id.title == CAR_NUMBER:
                car_number = attr.value
            if attr.pctransfer_attribute_id.title == CAR_DRIVER:
                car_driver = attr.value
        title = 'T A L A P N A M A     № ' + str(transfer.id) + ''
        date = '"' + str(current_day) + '"   ' + str(current_month) + '    ' + str(current_year) + 'ý.'
        to_whom_line = 'Kime (Кому)_____' + to_whom + '_______ Bölümi_____' + to_whom_dep + '____'
        from_whom_line = 'Kimden (От кого) ______' + from_whom + '_________ Bölümi ___' + from_whom_dep + '___'
        ws['A1']= title
        ws['A2']= date
        ws['A3']= to_whom_line
        ws['A4']= from_whom_line

        transfer_pallets = transfer_pallets.order_by('pallet_id__product_id__title')

        pallets = []
        p_counter = 0
        product = None

        for p in transfer_pallets:
            if p.pallet_id.product_id != product:
                pallets.append("")
            else:
                product = p.pallet_id.product_id
                p_counter = p_counter + 1

        product = None
        price_avg = 0
        total_amount  = 0
        total_price = 0
        p_counter = 0
        for p in transfer_pallets:
            unit = p.pallet_id.product_id.unit.title
            pallet_dict = {}
            if p.pallet_id.product_id.title == product:
                attributes = PalletPalletAttribute.objects.filter(pallet_id=p.pallet_id)
                for attr in attributes:
                    if attr.pallet_attr_id.title == unit:
                        total_amount = total_amount + attr.value
                        total_price = total_price + (p.price * (attr.value * 1.0))
                        price_avg = total_price / total_amount
                pallet_dict = {"code":p.pallet_id.container_id.id_number, 
                                "title":p.pallet_id.product_id.title, 
                                "unit":unit ,
                                "price":price_avg,
                                "total_amount":total_amount, 
                                "total_price":total_price}
                pallets[p_counter]=pallet_dict
                product = p.pallet_id.product_id.title
            else:
                p_counter = p_counter + 1
                total_amount  = 0
                total_price = 0
                price_avg = 0
                attributes = PalletPalletAttribute.objects.filter(pallet_id=p.pallet_id)
                for attr in attributes:
                    if attr.pallet_attr_id.title == unit:
                        total_amount = total_amount + attr.value
                        total_price = total_price + (p.price * (attr.value * 1.0))
                        price_avg = total_price / total_amount
                pallet_dict = {"code":p.pallet_id.container_id.id_number, 
                                "title":p.pallet_id.product_id.title, 
                                "unit":unit,
                                "price":price_avg,
                                "total_amount":total_amount, 
                                "total_price":total_price}
                pallets[p_counter]=pallet_dict
                product = p.pallet_id.product_id.title

        counter = 1
        for p in pallets:
            if p == '':
                continue
            row_number = counter + 7
            ws["A" + str(row_number)] = counter
            ws["B" + str(row_number)] = str(account_number)
            ws["C" + str(row_number)] = str(p["code"])
            ws["D" + str(row_number)] = str(p["title"])
            ws["E" + str(row_number)] = str(p["unit"])
            ws["F" + str(row_number)] = str(p["total_amount"])
            ws["G" + str(row_number)] = str(round(p["price"], 2))
            ws["H" + str(row_number)] = str(p["total_price"])

            counter = counter + 1
        
        ws['A18']= car_model
        ws['D18']= car_number
        ws['E18']= car_driver

        DIR_NAME = str(current_day) + str(month) + str(current_year)
        CHECK_FOLDER = os.path.isdir("media/files/" + DIR_NAME)

        if not CHECK_FOLDER:
            os.makedirs("media/files/" + DIR_NAME)
        wb.save('media/files/'+ DIR_NAME + "/talapnama_" + str(transfer.id) + '.xlsx')
        transfer.excel_file = "files/" + DIR_NAME + "/talapnama_" + str(transfer.id) + ".xlsx"
        transfer.save()
        ending = time.perf_counter()
        print(f"funtion run in {ending - starting} seconds.")
        return True
    except:
        return False


def product_report2(products):
    try:
        starting = time.perf_counter()
        template_file = 'xls_templates/harytlar.xlsx'
        wb = load_workbook(template_file)
        ws = wb.active
        today = datetime.date.today()
        day = today.day
        month = today.month
        year = today.year

        chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789123"
        char_length = len(chars)
        file_name = ""
        for i in range(12):
            file_name = file_name + chars[random.randrange(char_length)]

        title1 = '"Miweli ülke" hususy kärhanasynyň ammarlar toplumynyň '
        title2 = "ammar müdiri " + TEMP_STOCK_RESPONSIBLE + " jogapkärçiligindäki maddy gymmatlyklaryň"
        title3 = str(day) + "." + str(month) + "." + str(year) + " ý. ýagdaýyna "

        ws["A1"] = title1
        ws["A2"] = title2
        ws["A3"] = title3

        products_list = []
        for product in products:
            product_dict = {}
            total_amount = 0
            total_price = 0
            unit = ""
            if product.unit.title:
                unit = product.unit.title
            price_avg = 0
            cells = Cell.objects.filter(product_id=product)
            for cell in cells:
                pallet = cell.pallet_id
                transfer = PalletCellTransfer.objects.filter(pallet_id=pallet, transfer_id__transition_type=INCOME).first()
                if not transfer:
                    continue
                if pallet:
                    attributes = PalletPalletAttribute.objects.filter(pallet_id=pallet)
                    for attr in attributes:
                        if attr.pallet_attr_id.title == unit:
                            total_amount = total_amount + attr.value
                            total_price = total_price + (transfer.price * (attr.value * 1.0))
                            price_avg = total_price / total_amount
            product_dict = {
                "code": str(product.code),
                "title": str(product.title),
                "unit": unit,
                "price": round(price_avg, 2),
                "total_amount": total_amount,
                "total_price": total_price,
            }
            products_list.append(product_dict)    

        counter = 1
        total_sum = 0
        for product in products_list:
            row_number = counter + 6
            ws["A" + str(row_number)] = counter
            ws["C" + str(row_number)] = str(product["code"])
            ws["D" + str(row_number)] = str(product["title"])
            ws["E" + str(row_number)] = str(product["unit"])
            ws["F" + str(row_number)] = product["price"]
            ws["G" + str(row_number)] = product["total_amount"]
            ws["H" + str(row_number)] = product["total_price"]

            total_sum = total_sum + product["total_price"]

            counter = counter + 1

        total_sum = round(total_sum, 2)
        
        right_alignment = Alignment(horizontal="right", vertical="center")
        alignment = Alignment(horizontal="center", vertical="center")
        
        set_border(ws, 'A' + str(7) + ':H' + str(7+(counter - 1)))

        ws.merge_cells('A' + str(6+counter) + ':G' + str(6+counter))
        ws['A' + str(6+counter)] = "Jemi"
        ws['A' + str(6+counter)].alignment = right_alignment
        ws['A' + str(6+counter)].font = Font(bold=True, size=13)
        ws['H' + str(6+counter)] = total_sum


        ws.merge_cells('A' + str(9+counter) + ':C' + str(9+counter))
        ws.merge_cells('D' + str(9+counter) + ':E' + str(9+counter))
        ws.merge_cells('F' + str(9+counter) + ':H' + str(9+counter))
        ws.merge_cells('D' + str(10+counter) + ':E' + str(10+counter))
        ws['A' + str(9+counter)] = "Maddy jogapkär:"
        ws['A' + str(9+counter)].alignment = right_alignment
        ws['D' + str(9+counter)] = "____________________"
        ws['D' + str(9+counter)].alignment = alignment
        ws['F' + str(9+counter)] = TEMP_STOCK_RESPONSIBLE
        ws['F' + str(9+counter)].alignment = alignment
        ws['D' + str(10+counter)] = "goly"
        ws['D' + str(10+counter)].alignment = alignment
        ws['D' + str(10+counter)].font = Font(size=9)

        file_loc = "files/products/" + file_name + ".xlsx"
        wb.save("media/" + file_loc)
        ProductFiles.objects.create(file=file_loc)
        ending = time.perf_counter()
        print(f"funtion run in {ending - starting} seconds.")
        return file_loc
    except:
        return None

def talapnama_create2(pk):
        # starting = time.perf_counter()
    # try:
        title_font = Font(
            name='Times New Roman',
            size=24,
            bold=True,
        )
        number_font = Font(
            name="Times New Roman",
            size=18,
            bold=True,
            underline="single"
        )
        date_font = Font(
            name="Times New Roman",
            size=14,
        )
        
        content_font = Font(
            name="Times New Roman",
            size=12
        )
        alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
        border_bottom = Border(bottom=Side(style="thin"))

        transfer = PCTransfer.objects.get(pk=pk)
        transfer_created = transfer.created_at
        transfer_date = transfer_created.strftime("%d-%m-%Y")
        current_year = transfer_created.year
        month = transfer_created.month
        current_month = MONTHS[month]
        current_day = transfer_created.day

        transfer_pallets = PalletCellTransfer.objects.filter(
             transfer_id=transfer
        )
        transfer_attributes = TransferTransferAttribute.objects.filter(
             pctransfer_id = transfer
        )

        account_number = 0
        from_whom = ""
        to_whom = ""
        from_whom_dep = ""
        to_whom_dep = ""
        car_model = ""
        car_number = ""
        car_driver = ""
        for attr in transfer_attributes:
            if attr.pctransfer_attribute_id.title == HASAP_NO:
                account_number = attr.value
            if attr.pctransfer_attribute_id.title == TO_WHOM:
                to_whom = attr.value
            if attr.pctransfer_attribute_id.title == FROM_WHOM:
                from_whom = attr.value
            if attr.pctransfer_attribute_id.title == FROM_WHOM_DEP:
                from_whom_dep = attr.value
            if attr.pctransfer_attribute_id.title == TO_WHOM_DEP:
                to_whom_dep = attr.value
            if attr.pctransfer_attribute_id.title == CAR_MODEL:
                car_model = attr.value
            if attr.pctransfer_attribute_id.title == CAR_NUMBER:
                car_number = attr.value
            if attr.pctransfer_attribute_id.title == CAR_DRIVER:
                car_driver = attr.value
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Talapnama'
        ws.merge_cells('E1:H1')
        ws.merge_cells('A3:H3')

        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25

        alphabet = 'ABCDEFGHIJKLMN'
        for i in range(1,64):
            for j in alphabet:
                ws[str(j + str(i))].font = content_font
        
        title_row = ws.row_dimensions[1]
        title_row.height = 32
        title_cell = ws['D1']
        title_cell.value = "TALAPNAMA   "
        title_cell.alignment = right_alignment
        title_cell.font = title_font

        talapanama_number = ws["E1"]
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 32
        ws.column_dimensions["E"].width = 8
        talapanama_number.value = "№ " + str(transfer_date) + "/" + str(transfer.id)
        talapanama_number.font = number_font
        talapanama_number.alignment = left_alignment

        date_cell = ws['A3']
        date_cell.value = '"' + str(current_day) + '"   ' + str(current_month) + "  " + str(current_year) + "ý."
        date_cell.alignment = left_alignment
        date_cell.font = date_font

        to_whom_cell = ws["A5"]
        to_whom_cell.value = TO_WHOM
        from_whom_cell = ws["A7"]
        from_whom_cell.value = FROM_WHOM
        dep_to_whom_cell = ws["E5"]
        ws["C5"] = to_whom
        dep_from_whom_cell = ws["E7"]
        ws["C7"] = from_whom
        dep_to_whom_cell.value = "Bölümi"
        dep_from_whom_cell.value = "Bölümi"
        ws["F5"] = to_whom_dep
        ws["F7"] = from_whom_dep

        ws["A9"] = "T/N"
        ws["A9"].border = thin_border
        ws["A9"].alignment = alignment
        ws["B9"] = HASAP_NO
        ws["B9"].border = thin_border
        ws["B9"].alignment = alignment
        ws["C9"] = "Kody"
        ws["C9"].border = thin_border
        ws["C9"].alignment = alignment
        ws["D9"] = "Ady, hili, ölçegi"
        ws["D9"].border = thin_border
        ws["D9"].alignment = alignment
        ws["E9"] = "Ölçeg Birligi"
        ws["E9"].border = thin_border
        ws["E9"].alignment = alignment
        ws["E9"].alignment = Alignment(wrap_text=True)
        ws["F9"] = "Sany"
        ws["F9"].border = thin_border
        ws["F9"].alignment = alignment
        ws["G9"] = "Bahasy"
        ws["G9"].border = thin_border
        ws["G9"].alignment = alignment
        ws["H9"] = "Summasy"
        ws["H9"].border = thin_border
        ws["H9"].alignment = alignment

        
        pallets_count = len(transfer_pallets)
        counter = 1
        transfer_pallets = transfer_pallets.order_by('pallet_id__product_id__title')

        pallets = []
        p_counter = 0
        product = None

        for p in transfer_pallets:
            if p.pallet_id.product_id != product:
                pallets.append("")
            else:
                product = p.pallet_id.product_id
                p_counter = p_counter + 1

        product = None
        price_avg = 0
        total_amount  = 0
        total_price = 0
        p_counter = 0
        for p in transfer_pallets:
            unit = p.pallet_id.product_id.unit.title
            pallet_dict = {}
            if p.pallet_id.product_id.title == product:
                attributes = PalletPalletAttribute.objects.filter(pallet_id=p.pallet_id)
                for attr in attributes:
                    if attr.pallet_attr_id.title == unit:
                        total_amount = total_amount + attr.value
                        total_price = total_price + (p.price * (attr.value * 1.0))
                        price_avg = total_price / total_amount
                pallet_dict = {"code":p.pallet_id.container_id.id_number, 
                                "title":p.pallet_id.product_id.title, 
                                "unit":unit ,
                                "price":price_avg,
                                "total_amount":total_amount, 
                                "total_price":total_price}
                pallets[p_counter]=pallet_dict
                product = p.pallet_id.product_id.title
            else:
                p_counter = p_counter + 1
                total_amount  = 0
                total_price = 0
                price_avg = 0
                attributes = PalletPalletAttribute.objects.filter(pallet_id=p.pallet_id)
                for attr in attributes:
                    if attr.pallet_attr_id.title == unit:
                        total_amount = total_amount + attr.value
                        total_price = total_price + (p.price * (attr.value * 1.0))
                        price_avg = total_price / total_amount
                pallet_dict = {"code":p.pallet_id.container_id.id_number, 
                                "title":p.pallet_id.product_id.title, 
                                "unit":unit,
                                "price":price_avg,
                                "total_amount":total_amount, 
                                "total_price":total_price}
                pallets[p_counter]=pallet_dict
                product = p.pallet_id.product_id.title

        
        print(pallets)
        print(p_counter)

        pallets.remove('')
        

        counter = 1
        for p in pallets:
            if p == '':
                continue
            print(p)
            print("#######################")
            row_number = counter + 9
            ws["A" + str(row_number)] = counter
            ws["B" + str(row_number)] = str(account_number)
            counter = counter + 1
            ws["C" + str(row_number)] = str(p["code"])

            counter = counter + 1


        # counter = 1
        # for pallet in transfer_pallets:
        #     row_number = counter + 9
        #     ws.row_dimensions[row_number].height = 20
        #     ws["A" + str(row_number)] = counter
        #     ws["A" + str(row_number)].border = thin_border
        #     ws["B" + str(row_number)] = str(account_number)
        #     ws["B" + str(row_number)].border = thin_border
        #     ws["C" + str(row_number)] = str(pallet.pallet_id.code)
        #     ws["C" + str(row_number)].border = thin_border
        #     ws["D" + str(row_number)] = str(pallet.pallet_id.title) + " - " + str(pallet.pallet_id.product_id.title)
        #     ws["D" + str(row_number)].border = thin_border
        #     ws["D" + str(row_number)].alignment = Alignment(wrap_text=True)
        #     ws["E" + str(row_number)] = "" # Here should be unit
        #     ws["E" + str(row_number)].border = thin_border
        #     ws["F" + str(row_number)] = ""  # Here should be amount
        #     ws["F" + str(row_number)].border = thin_border
        #     ws["G" + str(row_number)] = pallet.price
        #     ws["G" + str(row_number)].border = thin_border
        #     ws["H" + str(row_number)] = pallet.price * 1  # Here , instaed of 1 should be amount
        #     ws["H" + str(row_number)].border = thin_border

        #     counter = counter + 1
        
        ws.merge_cells('A'+str(10+pallets_count)+":C"+str(10+pallets_count))
        ws.merge_cells('E'+str(10+pallets_count)+":H"+str(10+pallets_count))
        ws.merge_cells('A'+str(11+pallets_count)+":C"+str(11+pallets_count))
        ws.merge_cells('E'+str(11+pallets_count)+":H"+str(11+pallets_count))
        ws['A'+str(10+pallets_count)] = CAR_MODEL
        ws['A'+str(10+pallets_count)].alignment = alignment
        ws['A'+str(10+pallets_count)].border = thin_border
        ws['B'+str(10+pallets_count)].border = thin_border
        ws['C'+str(10+pallets_count)].border = thin_border        
        ws['D'+str(10+pallets_count)] = CAR_NUMBER
        ws['D'+str(10+pallets_count)].border = thin_border
        ws['D'+str(10+pallets_count)].alignment = alignment
        ws['E'+str(10+pallets_count)] = CAR_DRIVER
        ws['E'+str(10+pallets_count)].alignment = alignment
        ws['E'+str(10+pallets_count)].border = thin_border
        ws['F'+str(10+pallets_count)].border = thin_border
        ws['G'+str(10+pallets_count)].border = thin_border
        ws['H'+str(10+pallets_count)].border = thin_border

        ws['A'+str(11+pallets_count)] = car_model
        ws['A'+str(11+pallets_count)].border = thin_border
        ws['B'+str(11+pallets_count)].border = thin_border
        ws['C'+str(11+pallets_count)].border = thin_border
        ws['A'+str(11+pallets_count)].alignment = alignment
        ws['D'+str(11+pallets_count)] = car_number
        ws['D'+str(11+pallets_count)].border = thin_border
        ws['D'+str(11+pallets_count)].alignment = alignment
        ws['E'+str(11+pallets_count)] = car_driver
        ws['E'+str(11+pallets_count)].alignment = Alignment(wrap_text=True)
        ws['E'+str(11+pallets_count)].border = thin_border
        ws['F'+str(11+pallets_count)].border = thin_border
        ws['G'+str(11+pallets_count)].border = thin_border
        ws['H'+str(11+pallets_count)].border = thin_border
        ws['E'+str(11+pallets_count)].alignment = alignment

        ws["A" + str(13+pallets_count)] = DIRECTOR
        ws["C" + str(13+pallets_count)] = "________________"
        ws["E" + str(13+pallets_count)] = ACCOUNTANT
        ws["G" + str(13+pallets_count)] = "________________"

        ws["A" + str(15+pallets_count)] = SENT_BY
        ws["C" + str(15+pallets_count)] = "________________"
        ws["E" + str(15+pallets_count)] = RECIVED_BY
        ws["G" + str(15+pallets_count)] = "________________"

        ws.row_dimensions[10+pallets_count].height = 20
        ws.row_dimensions[11+pallets_count].height = 20
        
        DIR_NAME = str(current_day) + str(month) + str(current_year)
        CHECK_FOLDER = os.path.isdir("media/files/" + DIR_NAME)

        if not CHECK_FOLDER:
            os.makedirs("media/files/" + DIR_NAME)
        
        transfer.excel_file = "files/" + DIR_NAME + "/" + str(transfer.id) + ".xlsx"
        wb.save('media/files/'+ DIR_NAME + "/" + str(transfer.id) + '.xlsx')
        transfer.save()
        # ending = time.perf_counter()
        
        # print(f"funtion run in {ending - starting} seconds.")
        return True
    # except:
    #     return False
    
def product_report(products):
    content_font = Font(
            name="Times New Roman",
            size=12
        )
    chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789123"
    char_length = len(chars)
    file_name = ""
    for i in range(12):
        file_name = file_name + chars[random.randrange(char_length)]
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Harytlar'

    alphabet = 'ABCDEFGHIJKLMN'
    for i in range(1,64):
        for j in alphabet:
            ws[str(j + str(i))].font = content_font
    
    title_row = ws.row_dimensions[1]
    title_row.height = 32

    ws["A1"] = "№"
    ws["B1"] = "Kody"
    ws["C1"] = "Ady"
    ws["D1"] = "Ölçeg Birligi"
    ws["E1"] = "Möçberi"
    ws["F1"] = "Öýjükler"

    counter = 1
    for prod in products:
        ws["A" + str(counter + 1)] = counter
        ws["B" + str(counter + 1)] = prod["code"]
        ws["C" + str(counter + 1)] = prod["title"]
        ws["D" + str(counter + 1)] = prod["unit"]
        ws["E" + str(counter + 1)] = prod["amount"]
        ws["F" + str(counter + 1)] = ",".join(str(cell) for cell in prod["cells"])
        counter = counter + 1
    
    file_loc = "files/products/" + file_name + ".xlsx"
    wb.save("media/" + file_loc)
    # ProductFiles.objects.create(file=file_loc)
    return file_loc

def transfer_report(transfers):
    content_font = Font(
            name="Times New Roman",
            size=12
        )
    chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789123"
    char_length = len(chars)
    file_name = ""
    for i in range(12):
        file_name = file_name + chars[random.randrange(char_length)]
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hereketler'

    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    for i in range(1,64):
        for j in alphabet:
            ws[str(j + str(i))].font = content_font
    
    title_row = ws.row_dimensions[1]
    title_row.height = 32

    ws["A1"] = "№"
    ws["B1"] = "Görnüşi"
    ws["C1"] = "Sene"
    ws["D1"] = "Ulanyjy"
    ws["E1"] = FROM_WHOM
    ws["F1"] = FROM_WHOM_DEP
    ws["G1"] = TO_WHOM
    ws["H1"] = TO_WHOM_DEP
    ws["I1"] = CAR_MODEL
    ws["J1"] = CAR_NUMBER
    ws["K1"] = CAR_DRIVER
    ws["L1"] = 'Pallet'
    ws["M1"] = 'Haryt'
    ws["N1"] = 'Öýjuük'
    ws["O1"] = 'Baha'
    ws["P1"] = 'Ölçeg'
    ws["Q1"] = 'Möçberi'

    counter = 1
    for transfer in transfers:
        p_transfers = PalletCellTransfer.objects.filter(transfer_id=transfer['id'])
        
        for p_transfer in p_transfers:
            from_whom = ""
            to_whom = ""
            from_whom_dep = ""
            to_whom_dep = ""
            car_model = ""
            car_number = ""
            car_driver = ""
            unit = ""
            value = ""
            if transfer['attributes']:
                for t_attr in transfer['attributes']:
                    if t_attr['name'] == FROM_WHOM:
                        from_whom = t_attr['value']
                    if t_attr['name'] == TO_WHOM:
                        to_whom = t_attr['value']
                    if t_attr['name'] == FROM_WHOM_DEP:
                        from_whom_dep = t_attr['value']
                    if t_attr['name'] == TO_WHOM_DEP:
                        to_whom_dep = t_attr['value']
                    if t_attr['name'] == CAR_MODEL:
                        car_model = t_attr['value']
                    if t_attr['name'] == CAR_NUMBER:
                        car_number = t_attr['value']
                    if t_attr['name'] == CAR_DRIVER:
                        car_driver = t_attr['value']
            attrs = PalletPalletAttribute.objects.filter(pallet_id = p_transfer.pallet_id)
            for attr in attrs:
                print(attr.pallet_attr_id)
                if str(attr.pallet_attr_id.title) == str(p_transfer.pallet_id.product_id.unit):
                    unit = str(attr.pallet_attr_id.title)
                    value = str(attr.value)
                print("###############################")
            # for p_attr in p_transfer.pallet_id.at
            ws["A" + str(counter + 1)] = counter
            ws["B" + str(counter + 1)] = transfer["transition_type"]
            ws["c" + str(counter + 1)] = transfer["created_at"]
            ws["D" + str(counter + 1)] = transfer["user"]['username']
            ws["E" + str(counter + 1)] = str(from_whom)
            ws["F" + str(counter + 1)] = str(from_whom_dep)
            ws["G" + str(counter + 1)] = str(to_whom)
            ws["H" + str(counter + 1)] = str(to_whom_dep)
            ws["I" + str(counter + 1)] = str(car_model)
            ws["J" + str(counter + 1)] = str(car_number)
            ws["K" + str(counter + 1)] = str(car_driver)
            ws["L" + str(counter + 1)] = str(p_transfer.pallet_id.code) + " " + str(p_transfer.pallet_id.title)
            ws["M" + str(counter + 1)] = str(p_transfer.pallet_id.product_id.code) + " " + str(p_transfer.pallet_id.product_id.title)
            ws["N" + str(counter + 1)] = str(p_transfer.cell_id.code)
            ws["O" + str(counter + 1)] = str(p_transfer.price)
            ws["P" + str(counter + 1)] = str(unit)
            ws["Q" + str(counter + 1)] = str(value)
            counter = counter + 1
    file_loc = "files/transfers/" + file_name + ".xlsx"
    wb.save("media/" + file_loc)
    ProductFiles.objects.create(file=file_loc)
    return file_loc

def transfer_report2(startdate, enddate):
    starting = time.perf_counter()
    template_file = 'xls_templates/hereketler.xlsx'
    wb = load_workbook(template_file)

    ws = wb["Girdeji"]
    ws2 = wb["Çykdajy"]
    ws3 = wb["Hereket"]

    transfer_startdate = startdate
    transfer_enddate = enddate

    today = datetime.date.today()
    day = today.day
    month = today.month
    year = today.year

    if startdate and enddate:
        pctransfers_in = PCTransfer.objects.filter(created_at__range=[startdate, enddate],
                                                transition_type=INCOME)
        pctransfers_out = PCTransfer.objects.filter(created_at__range=[startdate, enddate],
                                                transition_type=OUTGO)
    else:
        startdate = "2023-01-01"
        enddate = today
        pctransfers_in = PCTransfer.objects.filter(transition_type=INCOME)
        pctransfers_out = PCTransfer.objects.filter(transition_type=OUTGO)


    title_in1 = '"Miweli ülke" hususy kärhanasynyň ammarlar toplumynyň ammar müdiri ' + TEMP_STOCK_RESPONSIBLE
    title_in2 =  str(startdate) + " - " + str(enddate) + " döwür aralygynda jogapkärçiligine girdeji bolan maddy gymmatlyklaryň"
    title_out1 = '"Miweli ülke" hususy kärhanasynyň ammarlar toplumynyň ammar müdiri ' + TEMP_STOCK_RESPONSIBLE
    title_out2 =  str(startdate) + " - " + str(enddate) + " döwür aralygynda jogapkärçiligine çykdajy bolan maddy gymmatlyklaryň"
    title_transfer1 = '"Miweli ülke" hususy kärhanasynyň ammarlar toplumynyň ammar müdiri ' + TEMP_STOCK_RESPONSIBLE
    title_transfer2 =  "jogapkärçiligindäki maddy gymmatlyklaryň " + str(startdate) + " - " + str(enddate) + " döwür aralygyndaky "

    ws["A1"] = title_in1
    ws["A2"] = title_in2
    ws2["A1"] = title_out1
    ws2["A2"] = title_out2
    ws3["A7"] = title_transfer1
    ws3["A8"] = title_transfer2

    
    counter = 1
    for transfer in pctransfers_in:
        transfer_pallets = PalletCellTransfer.objects.filter(
            transfer_id=transfer
        )
        transfer_attributes = TransferTransferAttribute.objects.filter(
                pctransfer_id = transfer
        )

        account_number = ""
        from_whom = ""
        to_whom = ""
        from_whom_dep = ""
        to_whom_dep = ""
        car_model = ""
        car_number = ""
        car_driver = ""
        for attr in transfer_attributes:
            if attr.pctransfer_attribute_id.title == HASAP_NO:
                account_number = attr.value
            if attr.pctransfer_attribute_id.title == TO_WHOM:
                to_whom = attr.value
            if attr.pctransfer_attribute_id.title == FROM_WHOM:
                from_whom = attr.value
            if attr.pctransfer_attribute_id.title == FROM_WHOM_DEP:
                from_whom_dep = attr.value
            if attr.pctransfer_attribute_id.title == TO_WHOM_DEP:
                to_whom_dep = attr.value
            if attr.pctransfer_attribute_id.title == CAR_MODEL:
                car_model = attr.value
            if attr.pctransfer_attribute_id.title == CAR_NUMBER:
                car_number = attr.value
            if attr.pctransfer_attribute_id.title == CAR_DRIVER:
                car_driver = attr.value
    
        transfer_pallets = transfer_pallets.order_by('pallet_id__product_id__title')
        pallets = []
        p_counter = 0
        product = None

        for p in transfer_pallets:
            if p.pallet_id.product_id != product:
                pallets.append("")
            else:
                product = p.pallet_id.product_id
                p_counter = p_counter + 1

        product = None
        price_avg = 0
        total_amount  = 0
        total_price = 0
        p_counter = 0
        for p in transfer_pallets:
            unit = p.pallet_id.product_id.unit.title
            pallet_dict = {}
            if p.pallet_id.product_id.title == product:
                attributes = PalletPalletAttribute.objects.filter(pallet_id=p.pallet_id)
                for attr in attributes:
                    if attr.pallet_attr_id.title == unit:
                        total_amount = total_amount + attr.value
                        total_price = total_price + (p.price * (attr.value * 1.0))
                        price_avg = total_price / total_amount
                pallet_dict = {"code":p.pallet_id.container_id.id_number, 
                                "title":p.pallet_id.product_id.title, 
                                "unit":unit ,
                                "price":price_avg,
                                "total_amount":total_amount, 
                                "total_price":total_price}
                pallets[p_counter]=pallet_dict
                product = p.pallet_id.product_id.title
            else:
                p_counter = p_counter + 1
                total_amount  = 0
                total_price = 0
                price_avg = 0
                attributes = PalletPalletAttribute.objects.filter(pallet_id=p.pallet_id)
                for attr in attributes:
                    if attr.pallet_attr_id.title == unit:
                        total_amount = total_amount + attr.value
                        total_price = total_price + (p.price * (attr.value * 1.0))
                        price_avg = total_price / total_amount
                pallet_dict = {"code":p.pallet_id.container_id.id_number, 
                                "title":p.pallet_id.product_id.title, 
                                "unit":unit,
                                "price":price_avg,
                                "total_amount":total_amount, 
                                "total_price":total_price}
                pallets[p_counter]=pallet_dict
                product = p.pallet_id.product_id.title

        for p in pallets:
            if p == '':
                continue
            row_number = counter + 6
            ws["A" + str(row_number)] = counter
            ws["B" + str(row_number)] = str(account_number)
            ws["C" + str(row_number)] = str(p["code"])
            ws["D" + str(row_number)] = str(p["title"])
            ws["E" + str(row_number)] = str(p["unit"])
            ws["F" + str(row_number)] = str(round(p["price"], 2))
            ws["G" + str(row_number)] = str(p["total_amount"])
            ws["H" + str(row_number)] = str(p["total_price"])
            ws["I" + str(row_number)] = transfer.batch_id.description
            ws["J" + str(row_number)] = ""
            ws["K" + str(row_number)] = ""
            ws["L" + str(row_number)] = car_number
            ws["M" + str(row_number)] = str(from_whom)
            ws["N" + str(row_number)] = ""
            ws["O" + str(row_number)] = ""

            counter = counter + 1
    

    counter = 1
    for transfer in pctransfers_out:
        transfer_pallets = PalletCellTransfer.objects.filter(
            transfer_id=transfer
        )
        transfer_attributes = TransferTransferAttribute.objects.filter(
                pctransfer_id = transfer
        )

        account_number = ""
        from_whom = ""
        to_whom = ""
        from_whom_dep = ""
        to_whom_dep = ""
        car_model = ""
        car_number = ""
        car_driver = ""
        for attr in transfer_attributes:
            if attr.pctransfer_attribute_id.title == HASAP_NO:
                account_number = attr.value
            if attr.pctransfer_attribute_id.title == TO_WHOM:
                to_whom = attr.value
            if attr.pctransfer_attribute_id.title == FROM_WHOM:
                from_whom = attr.value
            if attr.pctransfer_attribute_id.title == FROM_WHOM_DEP:
                from_whom_dep = attr.value
            if attr.pctransfer_attribute_id.title == TO_WHOM_DEP:
                to_whom_dep = attr.value
            if attr.pctransfer_attribute_id.title == CAR_MODEL:
                car_model = attr.value
            if attr.pctransfer_attribute_id.title == CAR_NUMBER:
                car_number = attr.value
            if attr.pctransfer_attribute_id.title == CAR_DRIVER:
                car_driver = attr.value
        
        transfer_pallets = transfer_pallets.order_by('pallet_id__product_id__title')
        pallets = []
        p_counter = 0
        product = None

        for p in transfer_pallets:
            if p.pallet_id.product_id != product:
                pallets.append("")
            else:
                product = p.pallet_id.product_id
                p_counter = p_counter + 1
        
        product = None
        price_avg = 0
        total_amount  = 0
        total_price = 0
        p_counter = 0

        for p in transfer_pallets:
            unit = p.pallet_id.product_id.unit.title
            pallet_dict = {}
            if p.pallet_id.product_id.title == product:
                attributes = PalletPalletAttribute.objects.filter(pallet_id=p.pallet_id)
                for attr in attributes:
                    if attr.pallet_attr_id.title == unit:
                        total_amount = total_amount + attr.value
                        total_price = total_price + (p.price * (attr.value * 1.0))
                        price_avg = total_price / total_amount
                pallet_dict = {"code":p.pallet_id.container_id.id_number, 
                                "title":p.pallet_id.product_id.title, 
                                "unit":unit,
                                "price":price_avg,
                                "total_amount":total_amount, 
                                "total_price":total_price}
                pallets[p_counter]=pallet_dict
                product = p.pallet_id.product_id.title
            else:
                p_counter = p_counter + 1
                total_amount  = 0
                total_price = 0
                price_avg = 0
                attributes = PalletPalletAttribute.objects.filter(pallet_id=p.pallet_id)
                for attr in attributes:
                    if attr.pallet_attr_id.title == unit:
                        total_amount = total_amount + attr.value
                        total_price = total_price + (p.price * (attr.value * 1.0))
                        price_avg = total_price / total_amount
                pallet_dict = {"code":p.pallet_id.container_id.id_number, 
                                "title":p.pallet_id.product_id.title, 
                                "unit":unit,
                                "price":price_avg,
                                "total_amount":total_amount, 
                                "total_price":total_price}
                pallets.append(pallet_dict)
                product = p.pallet_id.product_id.title
        
        for p in pallets:
            if p == '':
                continue
            row_number = counter + 6
            ws2["A" + str(row_number)] = counter
            ws2["B" + str(row_number)] = str(account_number)
            ws2["C" + str(row_number)] = str(p["code"])
            ws2["D" + str(row_number)] = str(p["title"])
            ws2["E" + str(row_number)] = str(p["unit"])
            ws2["F" + str(row_number)] = str(round(p["price"], 2))
            ws2["G" + str(row_number)] = str(p["total_amount"])
            ws2["H" + str(row_number)] = str(p["total_price"])
            ws2["I" + str(row_number)] = ""
            ws2["J" + str(row_number)] = ""
            ws2["K" + str(row_number)] = ""
            ws2["L" + str(row_number)] = car_number
            ws2["M" + str(row_number)] = str(from_whom)
            ws2["N" + str(row_number)] = ""
            ws2["O" + str(row_number)] = ""

            counter = counter + 1
    


    products = Product.objects.all()

    counter = 1
    for product in products:
        if transfer_startdate and transfer_enddate:
            product_log_start = ProductLog.objects.filter(product_id=product, created_at=transfer_startdate).first()
            product_log_end = ProductLog.objects.filter(product_id=product, created_at=transfer_enddate).first()
        else:
            transfer_startdate = "2023-01-01"
            transfer_enddate = today # - timedelta(days=1)
            product_log_start = ProductLog.objects.filter(product_id=product, created_at=transfer_startdate).first()
            product_log_end = ProductLog.objects.filter(product_id=product, created_at=transfer_enddate).first()
        
        row_counter = counter + 12

        pcctransfers_in = None
        for t_in in pctransfers_in:
            pcctransfers_in = PalletCellTransfer.objects.filter(transfer_id=t_in, pallet_id__product_id=product)
        
        pcctransfers_out = None
        for t_in in pctransfers_out:
            pcctransfers_out = PalletCellTransfer.objects.filter(transfer_id=t_in, pallet_id__product_id=product)
        


        total_sum_start = 0
        total_sum_end = 0
        total_amount_in  = 0
        total_price_in = 0
        total_amount_out = 0
        total_price_out = 0

        if pcctransfers_in:
            total_amount_in  = 0
            total_price_in = 0
            for p in pcctransfers_in:
                unit = product.unit.title
                attributes = PalletPalletAttribute.objects.filter(pallet_id=p.pallet_id)
                for attr in attributes:
                    if attr.pallet_attr_id.title == unit:
                        total_amount_in = total_amount_in + attr.value
                        total_price_in = total_price_in + (p.price * (attr.value * 1.0))

        
        if pcctransfers_out:
            total_amount_out = 0
            total_price_out = 0
            for p in pcctransfers_out:
                unit = product.unit.title
                attributes = PalletPalletAttribute.objects.filter(pallet_id=p.pallet_id)
                for attr in attributes:
                    if attr.pallet_attr_id.title == unit:
                        total_amount_out = total_amount_out + attr.value
                        total_price_out = total_price_out + (p.price * (attr.value * 1.0))



        ws3["A" + str(row_counter)] = counter
        ws3["C" + str(row_counter)] = str(product.code)
        ws3["D" + str(row_counter)] = str(product.title)
        ws3["E" + str(row_counter)] = str(product.unit.title)
        if product_log_start:
            if product_log_start.total_amount > 0:
                ws3["F" + str(row_counter)] = round(
                    (product_log_start.total_price / product_log_start.total_amount), 2)
            ws3["G" + str(row_counter)] = round(product_log_start.total_amount, 2)
            ws3["H" + str(row_counter)] = round(product_log_start.total_price, 2)
            total_sum_start = total_sum_start + product_log_start.total_price
        else:
            if product_log_end:
                if product_log_end.total_amount > 0:
                    ws3["F" + str(row_counter)] = round(
                        (product_log_end.total_price / product_log_end.total_amount), 2)
        if product_log_end:
            ws3["M" + str(row_counter)] = round(product_log_end.total_amount, 2)
            ws3["N" + str(row_counter)] = round(product_log_end.total_price, 2)
            total_sum_end = total_sum_end + product_log_end.total_price
        
        ws3["I" + str(row_counter)] = round(total_amount_in, 2)
        ws3["J" + str(row_counter)] = round(total_price_in, 2)
        ws3["K" + str(row_counter)] = round(total_amount_out, 2)
        ws3["L" + str(row_counter)] = round(total_price_out, 2)
        
        counter = counter + 1
    

    set_border(ws3, 'A' + str(13) + ':N' + str(13+(counter-1)))

    right_alignment = Alignment(horizontal="right", vertical="center")
    alignment = Alignment(horizontal="center", vertical="center")
    ws3.merge_cells('A' + str(12+counter) + ':F' + str(12+counter))
    ws3['A' + str(12+counter)] = "JEMI:"
    ws3['A' + str(12+counter)].alignment = right_alignment
    ws3['A' + str(12+counter)].font = Font(bold=True, size=13)
    ws3['H' + str(12+counter)] = "=SUM(H13:H" +  str(11+counter) + ")"
    ws3['J' + str(12+counter)] = "=SUM(J13:J" +  str(11+counter) + ")"
    ws3['L' + str(12+counter)] = "=SUM(L13:L" +  str(11+counter) + ")"

    ws3['N' + str(12+counter)] = "=SUM(N13:N" +  str(11+counter) + ")"
        

    file_name = create_name()
    file_loc = "files/transfers/" + file_name + ".xlsx"
    wb.save("media/" + file_loc)
    ending = time.perf_counter()
    print(f"funtion run in {ending - starting} seconds.")
    return file_loc